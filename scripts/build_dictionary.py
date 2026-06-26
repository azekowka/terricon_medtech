"""
Builds the normalized services dictionary for MedServicePrice.kz from the OFFICIAL
reference (data/source/services_reference.xlsx — the справочник услуг from the TZ).

The reference has columns: ID, Специальность, Code, Name_ru, TarificatrCode.
We map it into our dictionary schema and enrich it:

  * category   — derived from the tariff-code prefix (A=приём врача, B=лаборатория,
                 C=диагностика, D=процедура) with a name-based fallback.
  * specialty  — kept from the reference (122 specialties) as a richer sub-field.
  * tarif_code — kept for traceability.
  * synonyms   — auto-generated (parenthetical abbreviations, ё/приём variants) plus
                 our curated synonyms overlaid onto matching entries (rapidfuzz).
  * base_price_kzt / duration_days — heuristic price model by category/specialty,
                 overridden by curated prices where a curated service matches.

Output: data/services_dictionary.json

Run:  python scripts/build_dictionary.py
"""
import hashlib
import json
import os
import re

from openpyxl import load_workbook
from rapidfuzz import fuzz, process

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_XLSX = os.path.join(ROOT, "data", "source", "services_reference.xlsx")
CURATED = os.path.join(ROOT, "data", "source", "curated_overlay.json")
OUT = os.path.join(ROOT, "data", "services_dictionary.json")

CATEGORIES = {
    "laboratory": "Лаборатория",
    "doctor": "Приём врача",
    "diagnostic": "Диагностика",
    "procedure": "Процедура",
}
PREFIX_CATEGORY = {"A": "doctor", "B": "laboratory", "В": "laboratory", "C": "diagnostic", "D": "procedure"}

_DIAG_KW = ["узи", "мрт", "кт ", "рентген", "доплер", "томограф", "сцинтиграф",
            "маммограф", "флюорограф", "эхокг", "ктг", "холтер", "эндоскоп",
            "гастроскоп", "колоноскоп", "биопси", "денситометр"]
_DOC_KW = ["прием", "приём", "консультац", "осмотр"]
_LAB_KW = ["анализ", "кровь", "крови", "моча", "мочи", "мазок", "посев", "антитела",
           "igg", "igm", "пцр", "гормон", "маркер", "биохим", "соскоб"]


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").lower().replace("ё", "е")).strip()


def categorize(name: str, tarif: str) -> str:
    p = (tarif or "").strip()[:1]
    if p in PREFIX_CATEGORY:
        return PREFIX_CATEGORY[p]
    n = norm(name)
    if any(k in n for k in _DIAG_KW):
        return "diagnostic"
    if any(k in n for k in _DOC_KW):
        return "doctor"
    if any(k in n for k in _LAB_KW):
        return "laboratory"
    return "procedure"


def _jitter(name: str) -> float:
    return 0.82 + (int(hashlib.md5(name.encode()).hexdigest()[:6], 16) % 1000) / 1000 * 0.36


def base_price(name: str, specialty: str, category: str) -> int:
    n, sp = norm(name), norm(specialty)
    if category == "diagnostic":
        if "мрт" in n or "магнитно-резонанс" in n:
            base = 26000
        elif re.search(r"\bкт\b", n) or "компьютерная томограф" in n:
            base = 22000
        elif "ктг" in n or "кардиотоког" in n:
            base = 5000
        elif any(k in n for k in ("колоноскоп",)):
            base = 18000
        elif any(k in n for k in ("гастроскоп", "фгдс", "эндоскоп")):
            base = 12000
        elif "маммограф" in n:
            base = 6500
        elif "рентген" in n or "флюорограф" in n:
            base = 5000
        elif "узи" in n or "доплер" in n or "эхокг" in n:
            base = 7500
        else:
            base = 6000
    elif category == "doctor":
        base = 8500 if any(k in n for k in ("к.м.н", "профессор", "заведующ")) else 7000
    elif category == "laboratory":
        if "пцр" in sp or "пцр" in n:
            base = 8500
        elif "ифа" in sp:
            base = 4200
        elif "биохим" in sp:
            base = 1800
        elif "коагул" in sp:
            base = 3200
        elif "серолог" in sp:
            base = 3000
        elif "гормон" in sp or "гормон" in n:
            base = 3800
        elif "общая клиника" in sp:
            base = 2200
        else:
            base = 2500
    else:  # procedure
        base = 3500
    return int(round(base * _jitter(n) / 50) * 50)


def gen_synonyms(name: str) -> list[str]:
    syns: set[str] = set()
    m = re.search(r"\(([^)]{2,25})\)", name)
    if m:
        syns.add(m.group(1).strip())
    stripped = re.sub(r"\s*\([^)]*\)", "", name).strip()
    if stripped and stripped != name:
        syns.add(stripped)
    if name.startswith("Прием"):
        syns.add(name.replace("Прием", "Приём", 1))
        syns.add(name.replace("Прием ", "Консультация ", 1))
    return [s for s in syns if s and norm(s) != norm(name)]


def load_curated():
    """normalized curated name -> (name, synonyms, base_price, duration_days)."""
    with open(CURATED, encoding="utf-8") as f:
        items = json.load(f)
    return [
        {
            "name": it["name"],
            "synonyms": it.get("synonyms", []),
            "base_price_kzt": it.get("base_price_kzt"),
            "duration_days": it.get("duration_days"),
        }
        for it in items
    ]


def build():
    wb = load_workbook(SRC_XLSX, data_only=True)
    ws = wb.active
    rows = [r for r in ws.iter_rows(values_only=True)][1:]
    rows = [r for r in rows if r and r[3]]

    def _int(v):
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    # de-dup by normalized name (keep first occurrence)
    seen: set[str] = set()
    entries = []
    for idx, (offid, specialty, code, name_ru, tarif) in enumerate(rows):
        name = str(name_ru).strip()
        nn = norm(name)
        if nn in seen:
            continue
        seen.add(nn)
        category = categorize(name, tarif)
        cid = _int(code)
        entries.append({
            "code": f"svc_{cid}" if cid is not None else f"svc_x{idx}",
            "official_id": cid,
            "name": name,
            "specialty": str(specialty or "").strip(),
            "tarif_code": str(tarif or "").strip(),
            "category": category,
            "category_label": CATEGORIES[category],
            "synonyms": gen_synonyms(name),
            "base_price_kzt": base_price(name, specialty or "", category),
            "duration_days": 1 if category == "laboratory" else None,
            "is_core": False,  # set True for curated common services below
        })

    # overlay curated synonyms + prices via fuzzy name match
    by_norm = {norm(e["name"]): e for e in entries}
    keys = list(by_norm.keys())
    overlaid = 0
    for cur in load_curated():
        ck = norm(cur["name"])
        target = by_norm.get(ck)
        if target is None:
            best = process.extractOne(ck, keys, scorer=fuzz.WRatio)
            if best and best[1] >= 90:
                target = by_norm[best[0]]
        if target is None:
            continue
        merged = list(dict.fromkeys([*target["synonyms"], *cur["synonyms"]]))
        target["synonyms"] = merged
        if cur.get("base_price_kzt"):
            target["base_price_kzt"] = cur["base_price_kzt"]
        if cur.get("duration_days") is not None:
            target["duration_days"] = cur["duration_days"]
        target["is_core"] = True  # curated -> common service offered widely in seed
        overlaid += 1

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

    by_cat: dict[str, int] = {}
    for e in entries:
        by_cat[e["category"]] = by_cat.get(e["category"], 0) + 1
    n_spec = len({e["specialty"] for e in entries})
    print(f"Wrote {len(entries)} services (official reference) to {OUT}")
    print("By category:", by_cat)
    print(f"Specialties: {n_spec} | curated overlays applied: {overlaid}")


if __name__ == "__main__":
    build()
